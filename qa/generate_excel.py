#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generateur du Guide de Tests Techniques / Fonctionnels - Mood-IoT
Format Excel (.xlsx) avec openpyxl - Tout en francais
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE_DIR, "Guion_de_Pruebas_MoodIoT.xlsx")

# Couleurs
BLEU_FONCE = "1F4E79"
BLEU_MOYEN = "2E75B6"
BLEU_CLAIR = "D6E4F0"
VERT = "C6EFCE"
ROUGE = "FFC7CE"
JAUNE = "FFEB9C"
BLANC = "FFFFFF"
GRIS_CLAIR = "F2F2F2"

# Styles communs
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
FONT_HEADER = Font(name="Calibri", bold=True, color=BLANC, size=10)
FONT_BODY = Font(name="Calibri", size=10)
FONT_TITLE = Font(name="Calibri", bold=True, size=14, color=BLEU_FONCE)
FILL_HEADER = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
FILL_ALT = PatternFill(start_color=GRIS_CLAIR, end_color=GRIS_CLAIR, fill_type="solid")
FILL_GREEN = PatternFill(start_color=VERT, end_color=VERT, fill_type="solid")
FILL_RED = PatternFill(start_color=ROUGE, end_color=ROUGE, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=JAUNE, end_color=JAUNE, fill_type="solid")
FILL_BLEU_CLAIR = PatternFill(start_color=BLEU_CLAIR, end_color=BLEU_CLAIR, fill_type="solid")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ---------------------------------------------------------------------------
# 52 cas de tests
# ---------------------------------------------------------------------------
TEST_CASES = [
    # UC 1 : Infrastructure et Deploiement (3 cas)
    {
        "req": "UC 1", "segment": "Infrastructure et Deploiement",
        "id": "TC-UC1-01",
        "nom": "Verification du demarrage des 10 conteneurs Docker",
        "criticite": "Haute",
        "prerequis": "Docker Compose installe, fichier docker-compose.yml present",
        "etapes": "1. Executer docker compose up -d\n2. Executer docker compose ps\n3. Verifier que 10 conteneurs sont en etat running",
        "resultat": "10 conteneurs affiches avec statut running/healthy",
        "etat": "Reussi",
        "observations": "Tous les conteneurs demarrent en moins de 30 secondes",
        "solution": "-",
        "evidences": "01_docker_ps.txt",
    },
    {
        "req": "UC 1", "segment": "Infrastructure et Deploiement",
        "id": "TC-UC1-02",
        "nom": "PostgreSQL avec 17 tables et donnees initiales",
        "criticite": "Haute",
        "prerequis": "Conteneur PostgreSQL demarre",
        "etapes": "1. Se connecter a la base de donnees PostgreSQL\n2. Lister les tables avec \\dt\n3. Verifier la presence des tables users et patients",
        "resultat": "17 tables listees incluant users et patients avec donnees de seed",
        "etat": "Reussi",
        "observations": "Schema conforme au modele de donnees",
        "solution": "-",
        "evidences": "02_db_tables.txt",
    },
    {
        "req": "UC 1", "segment": "Infrastructure et Deploiement",
        "id": "TC-UC1-03",
        "nom": "Redis connecte et operationnel",
        "criticite": "Moyenne",
        "prerequis": "Conteneur Redis demarre",
        "etapes": "1. Verifier le conteneur Redis dans docker compose ps\n2. Tester la connectivite Redis",
        "resultat": "Redis est actif et repond aux commandes PING",
        "etat": "Reussi",
        "observations": "Redis utilise pour le cache des sessions",
        "solution": "-",
        "evidences": "01_docker_ps.txt",
    },
    # UC 2 : Verifications de sante (6 cas)
    {
        "req": "UC 2", "segment": "Verifications de sante",
        "id": "TC-UC2-01",
        "nom": "Health check Gateway agrege",
        "criticite": "Haute",
        "prerequis": "Gateway demarre sur le port 4000",
        "etapes": "1. Envoyer GET http://localhost:4000/health\n2. Verifier la reponse JSON avec status: ok",
        "resultat": "Reponse 200 OK avec status: ok et details des services",
        "etat": "Reussi",
        "observations": "Tous les services sous-jacents signales comme sains",
        "solution": "-",
        "evidences": "03_gateway_health.json",
    },
    {
        "req": "UC 2", "segment": "Verifications de sante",
        "id": "TC-UC2-02",
        "nom": "Health check direct service Auth",
        "criticite": "Haute",
        "prerequis": "Service Auth demarre sur le port 4001",
        "etapes": "1. Envoyer GET http://localhost:4001/health\n2. Verifier la reponse JSON",
        "resultat": "Reponse 200 OK avec status: ok",
        "etat": "Reussi",
        "observations": "Service Auth operationnel",
        "solution": "-",
        "evidences": "04_auth_health.json",
    },
    {
        "req": "UC 2", "segment": "Verifications de sante",
        "id": "TC-UC2-03",
        "nom": "Health check direct service Patient",
        "criticite": "Haute",
        "prerequis": "Service Patient demarre sur le port 4002",
        "etapes": "1. Envoyer GET http://localhost:4002/health\n2. Verifier la reponse JSON",
        "resultat": "Reponse 200 OK avec status: ok",
        "etat": "Reussi",
        "observations": "Service Patient operationnel",
        "solution": "-",
        "evidences": "05_patient_health.json",
    },
    {
        "req": "UC 2", "segment": "Verifications de sante",
        "id": "TC-UC2-04",
        "nom": "Health check direct service Scoring",
        "criticite": "Haute",
        "prerequis": "Service Scoring demarre sur le port 4003",
        "etapes": "1. Envoyer GET http://localhost:4003/health\n2. Verifier la reponse JSON",
        "resultat": "Reponse 200 OK avec status: ok",
        "etat": "Reussi",
        "observations": "Service Scoring operationnel",
        "solution": "-",
        "evidences": "06_scoring_health.json",
    },
    {
        "req": "UC 2", "segment": "Verifications de sante",
        "id": "TC-UC2-05",
        "nom": "Health check direct service Notification",
        "criticite": "Moyenne",
        "prerequis": "Service Notification demarre sur le port 4004",
        "etapes": "1. Envoyer GET http://localhost:4004/health\n2. Verifier la reponse JSON",
        "resultat": "Reponse 200 OK avec status: ok",
        "etat": "Reussi",
        "observations": "Service Notification operationnel",
        "solution": "-",
        "evidences": "07_notif_health.json",
    },
    {
        "req": "UC 2", "segment": "Verifications de sante",
        "id": "TC-UC2-06",
        "nom": "Health check direct service Teleconsultation",
        "criticite": "Moyenne",
        "prerequis": "Service Teleconsultation demarre sur le port 4005",
        "etapes": "1. Envoyer GET http://localhost:4005/health\n2. Verifier la reponse JSON",
        "resultat": "Reponse 200 OK avec status: ok",
        "etat": "Reussi",
        "observations": "Service Teleconsultation operationnel",
        "solution": "-",
        "evidences": "08_teleconsult_health.json",
    },
    # UC 3 : Authentification et Sessions (13 cas)
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-01",
        "nom": "Connexion reussie psychiatre (JWT)",
        "criticite": "Haute",
        "prerequis": "Service Auth demarre, utilisateur psychiatre cree en base",
        "etapes": "1. POST /auth/login avec identifiants psychiatre\n2. Verifier le code 200\n3. Verifier la presence du token JWT dans la reponse",
        "resultat": "Reponse 200 avec access_token, refresh_token et role psychiatre",
        "etat": "Reussi",
        "observations": "Token JWT valide avec expiration configuree",
        "solution": "-",
        "evidences": "09_auth_login_psychiatre.json, sc_03_dashboard_apres_login.png",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-02",
        "nom": "Connexion reussie patient Sophie",
        "criticite": "Haute",
        "prerequis": "Service Auth demarre, utilisateur patient cree en base",
        "etapes": "1. POST /auth/login avec identifiants patient Sophie\n2. Verifier le code 200\n3. Verifier la presence du token JWT",
        "resultat": "Reponse 200 avec access_token et role patient",
        "etat": "Reussi",
        "observations": "Patient Sophie authentifie avec succes",
        "solution": "-",
        "evidences": "10_auth_login_patient.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-03",
        "nom": "Connexion mot de passe incorrect - erreur 401",
        "criticite": "Haute",
        "prerequis": "Service Auth demarre",
        "etapes": "1. POST /auth/login avec mot de passe incorrect\n2. Verifier le code 401\n3. Verifier le message d'erreur",
        "resultat": "Reponse 401 Unauthorized avec message identifiants invalides",
        "etat": "Reussi",
        "observations": "Message d'erreur generique pour eviter l'enumeration",
        "solution": "-",
        "evidences": "11_auth_login_wrong.json, sc_02_login_erreur.png",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-04",
        "nom": "GET /auth/me psychiatre",
        "criticite": "Haute",
        "prerequis": "Token JWT valide du psychiatre",
        "etapes": "1. GET /auth/me avec Authorization Bearer token\n2. Verifier le code 200\n3. Verifier les donnees utilisateur",
        "resultat": "Reponse 200 avec profil complet du psychiatre",
        "etat": "Reussi",
        "observations": "Profil retourne avec role et permissions",
        "solution": "-",
        "evidences": "12_auth_me_psychiatre.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-05",
        "nom": "GET /auth/me patient",
        "criticite": "Haute",
        "prerequis": "Token JWT valide du patient",
        "etapes": "1. GET /auth/me avec Authorization Bearer token patient\n2. Verifier le code 200",
        "resultat": "Reponse 200 avec profil du patient",
        "etat": "Reussi",
        "observations": "Profil patient retourne correctement",
        "solution": "-",
        "evidences": "13_auth_me_patient.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-06",
        "nom": "GET /auth/me sans token - erreur 403",
        "criticite": "Haute",
        "prerequis": "Aucun token fourni",
        "etapes": "1. GET /auth/me sans header Authorization\n2. Verifier le code 403",
        "resultat": "Reponse 403 Forbidden",
        "etat": "Reussi",
        "observations": "Acces refuse sans authentification",
        "solution": "-",
        "evidences": "14_auth_me_notoken.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-07",
        "nom": "Rafraichissement du token JWT",
        "criticite": "Haute",
        "prerequis": "Refresh token valide obtenu lors de la connexion",
        "etapes": "1. POST /auth/refresh avec le refresh_token\n2. Verifier le code 200\n3. Verifier le nouveau access_token",
        "resultat": "Nouveau access_token genere avec succes",
        "etat": "Reussi",
        "observations": "Rotation des tokens fonctionnelle",
        "solution": "-",
        "evidences": "15_auth_refresh.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-08",
        "nom": "Configuration MFA TOTP",
        "criticite": "Moyenne",
        "prerequis": "Utilisateur authentifie",
        "etapes": "1. POST /auth/mfa/setup avec token valide\n2. Verifier la reponse avec secret TOTP et QR code",
        "resultat": "Secret TOTP et URI otpauth generes",
        "etat": "Reussi",
        "observations": "QR code generee pour configuration dans l'application",
        "solution": "-",
        "evidences": "16_auth_mfa_setup.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-09",
        "nom": "Inscription nouvel utilisateur",
        "criticite": "Haute",
        "prerequis": "Service Auth demarre",
        "etapes": "1. POST /auth/register avec donnees du nouvel utilisateur\n2. Verifier le code 201\n3. Verifier la creation du compte",
        "resultat": "Compte cree avec succes, reponse 201 avec donnees utilisateur",
        "etat": "Reussi",
        "observations": "Mot de passe hache en base de donnees",
        "solution": "-",
        "evidences": "17_auth_register.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-10",
        "nom": "Inscription email duplique - erreur 409",
        "criticite": "Haute",
        "prerequis": "Email deja enregistre en base",
        "etapes": "1. POST /auth/register avec email existant\n2. Verifier le code 409",
        "resultat": "Reponse 409 Conflict indiquant que l'email est deja utilise",
        "etat": "Reussi",
        "observations": "Contrainte d'unicite respectee",
        "solution": "-",
        "evidences": "18_auth_register_dup.json",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-11",
        "nom": "Deconnexion invalide le token",
        "criticite": "Moyenne",
        "prerequis": "Token JWT valide",
        "etapes": "1. POST /auth/logout avec token valide\n2. Verifier le code 200\n3. Tenter GET /auth/me avec l'ancien token",
        "resultat": "Deconnexion reussie, ancien token invalide",
        "etat": "Reussi",
        "observations": "Token ajoute a la liste noire Redis",
        "solution": "-",
        "evidences": "19_auth_logout.json, sc_08_deconnexion.png",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-12",
        "nom": "Stockage token JWT dans localStorage",
        "criticite": "Haute",
        "prerequis": "Dashboard Next.js charge dans le navigateur",
        "etapes": "1. Se connecter via le formulaire du dashboard\n2. Ouvrir les DevTools > Application > localStorage\n3. Verifier les cles mood_token, mood_refresh, mood_user",
        "resultat": "Tokens stockes dans localStorage sous les cles mood_token, mood_refresh et mood_user",
        "etat": "Reussi",
        "observations": "Tokens persistes cote client pour maintenir la session",
        "solution": "-",
        "evidences": "Verification manuelle : mood_token, mood_refresh, mood_user presents dans localStorage",
    },
    {
        "req": "UC 3", "segment": "Authentification et Sessions",
        "id": "TC-UC3-13",
        "nom": "Page de connexion du dashboard",
        "criticite": "Haute",
        "prerequis": "Dashboard Next.js accessible sur http://localhost:3000",
        "etapes": "1. Acceder a http://localhost:3000\n2. Verifier l'affichage du formulaire de connexion\n3. Verifier les champs email et mot de passe",
        "resultat": "Page de connexion affichee avec formulaire fonctionnel",
        "etat": "Reussi",
        "observations": "Design responsive et conforme a la maquette",
        "solution": "-",
        "evidences": "sc_01_login_page.png",
    },
    # UC 4 : Service Patient (PostgreSQL) (7 cas)
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-01",
        "nom": "Lister 4 patients depuis PostgreSQL",
        "criticite": "Haute",
        "prerequis": "Service Patient demarre, donnees de seed chargees",
        "etapes": "1. GET /patients avec token psychiatre\n2. Verifier le code 200\n3. Verifier que 4 patients sont retournes",
        "resultat": "Liste de 4 patients avec donnees completes",
        "etat": "Reussi",
        "observations": "Pagination disponible si necessaire",
        "solution": "-",
        "evidences": "20_patient_list.json",
    },
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-02",
        "nom": "Obtenir patient par UUID",
        "criticite": "Haute",
        "prerequis": "UUID d'un patient connu",
        "etapes": "1. GET /patients/{uuid} avec token\n2. Verifier le code 200\n3. Verifier les donnees du patient",
        "resultat": "Donnees completes du patient retournees",
        "etat": "Reussi",
        "observations": "UUID conforme au format standard",
        "solution": "-",
        "evidences": "21_patient_get.json",
    },
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-03",
        "nom": "Soumission PHQ-9 et calcul severite",
        "criticite": "Haute",
        "prerequis": "Patient authentifie",
        "etapes": "1. POST /patients/{id}/mood avec donnees PHQ-9\n2. Verifier le code 201\n3. Verifier le score et la severite calcules",
        "resultat": "Score PHQ-9 calcule avec niveau de severite correspondant",
        "etat": "Reussi",
        "observations": "Algorithme de scoring conforme aux references cliniques",
        "solution": "-",
        "evidences": "22_patient_mood_phq9.json",
    },
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-04",
        "nom": "Sync donnees sante Health Connect (UPSERT)",
        "criticite": "Haute",
        "prerequis": "Patient authentifie, donnees Health Connect disponibles",
        "etapes": "1. POST /patients/{id}/healthdata avec donnees Health Connect\n2. Verifier le code 200/201\n3. Verifier l'operation UPSERT",
        "resultat": "Donnees de sante synchronisees avec succes (insertion ou mise a jour)",
        "etat": "Reussi",
        "observations": "UPSERT evite les doublons de donnees",
        "solution": "-",
        "evidences": "23_patient_healthdata.json",
    },
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-05",
        "nom": "Sync batch donnees sante (2 jours)",
        "criticite": "Haute",
        "prerequis": "Patient authentifie",
        "etapes": "1. POST /patients/{id}/healthdata/batch avec 2 jours de donnees\n2. Verifier le code 200\n3. Verifier le traitement batch",
        "resultat": "2 enregistrements journaliers traites avec succes",
        "etat": "Reussi",
        "observations": "Traitement batch optimise pour les synchronisations volumineuses",
        "solution": "-",
        "evidences": "24_patient_batch.json",
    },
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-06",
        "nom": "Plateforme invalide - erreur 422",
        "criticite": "Moyenne",
        "prerequis": "Patient authentifie",
        "etapes": "1. POST /patients/{id}/healthdata avec plateforme invalide\n2. Verifier le code 422",
        "resultat": "Reponse 422 Unprocessable Entity avec details de validation",
        "etat": "Reussi",
        "observations": "Validation stricte des plateformes supportees",
        "solution": "-",
        "evidences": "25_patient_invalid_platform.json",
    },
    {
        "req": "UC 4", "segment": "Service Patient (PostgreSQL)",
        "id": "TC-UC4-07",
        "nom": "Consulter consentements PostgreSQL",
        "criticite": "Moyenne",
        "prerequis": "Patient avec consentements enregistres",
        "etapes": "1. GET /patients/{id}/consents avec token\n2. Verifier le code 200\n3. Verifier la liste des consentements",
        "resultat": "Liste des consentements du patient retournee",
        "etat": "Reussi",
        "observations": "Conformite RGPD pour la gestion des consentements",
        "solution": "-",
        "evidences": "26_patient_consents.json",
    },
    # UC 5 : Service Teleconsultation (PostgreSQL) (2 cas)
    {
        "req": "UC 5", "segment": "Service Teleconsultation (PostgreSQL)",
        "id": "TC-UC5-01",
        "nom": "Creer session teleconsultation avec Jitsi",
        "criticite": "Haute",
        "prerequis": "Psychiatre authentifie, patient existant",
        "etapes": "1. POST /teleconsult/sessions avec donnees de la session\n2. Verifier le code 201\n3. Verifier l'URL Jitsi generee",
        "resultat": "Session creee avec URL Jitsi unique",
        "etat": "Reussi",
        "observations": "Integration Jitsi Meet fonctionnelle",
        "solution": "-",
        "evidences": "27_teleconsult_create.json",
    },
    {
        "req": "UC 5", "segment": "Service Teleconsultation (PostgreSQL)",
        "id": "TC-UC5-02",
        "nom": "Lister sessions teleconsultation",
        "criticite": "Moyenne",
        "prerequis": "Sessions de teleconsultation existantes",
        "etapes": "1. GET /teleconsult/sessions avec token\n2. Verifier le code 200\n3. Verifier la liste des sessions",
        "resultat": "Liste des sessions retournee avec details",
        "etat": "Reussi",
        "observations": "Filtrage par psychiatre/patient disponible",
        "solution": "-",
        "evidences": "28_teleconsult_list.json",
    },
    # UC 6 : Service Scoring (1 cas)
    {
        "req": "UC 6", "segment": "Service Scoring",
        "id": "TC-UC6-01",
        "nom": "Historique des scores d'un patient",
        "criticite": "Moyenne",
        "prerequis": "Patient avec historique de scores",
        "etapes": "1. GET /scoring/patients/{id}/history avec token\n2. Verifier le code 200\n3. Verifier l'historique des scores",
        "resultat": "Historique des scores retourne avec dates et valeurs",
        "etat": "Reussi",
        "observations": "Scores tries par date decroissante",
        "solution": "-",
        "evidences": "29_scoring_history.json",
    },
    # UC 7 : Proxy Gateway (3 cas)
    {
        "req": "UC 7", "segment": "Proxy Gateway",
        "id": "TC-UC7-01",
        "nom": "Connexion via Gateway proxy",
        "criticite": "Haute",
        "prerequis": "Gateway demarre sur le port 4000",
        "etapes": "1. POST http://localhost:4000/auth/login\n2. Verifier le routage vers le service Auth\n3. Verifier la reponse",
        "resultat": "Connexion reussie via le proxy Gateway",
        "etat": "Reussi",
        "observations": "Gateway route correctement vers le service Auth",
        "solution": "-",
        "evidences": "30_gw_auth_login.json",
    },
    {
        "req": "UC 7", "segment": "Proxy Gateway",
        "id": "TC-UC7-02",
        "nom": "/auth/me via Gateway",
        "criticite": "Haute",
        "prerequis": "Token JWT valide, Gateway demarre",
        "etapes": "1. GET http://localhost:4000/auth/me avec token\n2. Verifier le routage et la reponse",
        "resultat": "Profil utilisateur retourne via Gateway",
        "etat": "Reussi",
        "observations": "Headers de proxy correctement transmis",
        "solution": "-",
        "evidences": "31_gw_auth_me.json",
    },
    {
        "req": "UC 7", "segment": "Proxy Gateway",
        "id": "TC-UC7-03",
        "nom": "Sante scoring via Gateway",
        "criticite": "Moyenne",
        "prerequis": "Gateway et service Scoring demarres",
        "etapes": "1. GET http://localhost:4000/scoring/health\n2. Verifier la reponse",
        "resultat": "Health check scoring accessible via Gateway",
        "etat": "Reussi",
        "observations": "Routage Gateway vers Scoring fonctionnel",
        "solution": "-",
        "evidences": "32_gw_scoring_health.json",
    },
    # UC 8 : Persistance en base de donnees (1 cas)
    {
        "req": "UC 8", "segment": "Persistance en base de donnees",
        "id": "TC-UC8-01",
        "nom": "Donnees persistees apres operations CRUD",
        "criticite": "Haute",
        "prerequis": "Operations CRUD effectuees sur la base",
        "etapes": "1. Effectuer des operations CRUD via l'API\n2. Verifier le comptage des enregistrements en base\n3. Confirmer la persistance",
        "resultat": "Nombre d'enregistrements coherent apres les operations",
        "etat": "Reussi",
        "observations": "Transactions ACID respectees",
        "solution": "-",
        "evidences": "33_db_persistence.txt",
    },
    # UC 9 : Dashboard Frontend Next.js (6 cas)
    {
        "req": "UC 9", "segment": "Dashboard Frontend Next.js",
        "id": "TC-UC9-01",
        "nom": "Page de connexion medecin",
        "criticite": "Haute",
        "prerequis": "Dashboard Next.js accessible sur http://localhost:3000",
        "etapes": "1. Acceder a http://localhost:3000\n2. Verifier le formulaire de connexion\n3. Verifier la presence des champs email et mot de passe",
        "resultat": "Formulaire de connexion affiche avec validation cote client",
        "etat": "Reussi",
        "observations": "Interface conforme aux maquettes Figma",
        "solution": "-",
        "evidences": "sc_01_login_page.png",
    },
    {
        "req": "UC 9", "segment": "Dashboard Frontend Next.js",
        "id": "TC-UC9-02",
        "nom": "Dashboard principal avec 4 KPIs",
        "criticite": "Haute",
        "prerequis": "Medecin connecte au dashboard",
        "etapes": "1. Se connecter avec les identifiants du psychiatre\n2. Verifier l'affichage du dashboard principal\n3. Verifier les 4 indicateurs KPI",
        "resultat": "Dashboard affiche avec 4 KPIs : patients actifs, alertes, teleconsultations, score moyen",
        "etat": "Reussi",
        "observations": "KPIs calcules en temps reel depuis l'API",
        "solution": "-",
        "evidences": "sc_04_dashboard_complet.png",
    },
    {
        "req": "UC 9", "segment": "Dashboard Frontend Next.js",
        "id": "TC-UC9-03",
        "nom": "Fiche patiente avec metriques vs baseline",
        "criticite": "Haute",
        "prerequis": "Patient existant avec historique de donnees",
        "etapes": "1. Cliquer sur un patient dans la liste\n2. Verifier l'affichage de la fiche patiente\n3. Comparer les metriques actuelles avec le baseline",
        "resultat": "Fiche patiente affichee avec comparaison metriques vs baseline",
        "etat": "Reussi",
        "observations": "Graphiques de tendance visibles",
        "solution": "-",
        "evidences": "sc_05_fiche_patiente.png",
    },
    {
        "req": "UC 9", "segment": "Dashboard Frontend Next.js",
        "id": "TC-UC9-04",
        "nom": "Page notifications avec badge",
        "criticite": "Moyenne",
        "prerequis": "Notifications existantes pour le psychiatre",
        "etapes": "1. Cliquer sur l'icone de notifications\n2. Verifier l'affichage des alertes\n3. Verifier le badge de compteur",
        "resultat": "Liste des notifications affichee avec badge de compteur",
        "etat": "Reussi",
        "observations": "Alertes classees par priorite",
        "solution": "-",
        "evidences": "sc_06_notifications.png",
    },
    {
        "req": "UC 9", "segment": "Dashboard Frontend Next.js",
        "id": "TC-UC9-05",
        "nom": "Messagerie avec messages rapides",
        "criticite": "Moyenne",
        "prerequis": "Medecin connecte au dashboard",
        "etapes": "1. Acceder a la page messagerie\n2. Verifier l'affichage des conversations\n3. Tester les boutons de messages rapides",
        "resultat": "Messagerie fonctionnelle avec boutons de reponses rapides",
        "etat": "Reussi",
        "observations": "Messages rapides predefinis pour gain de temps",
        "solution": "-",
        "evidences": "sc_07_messagerie.png",
    },
    {
        "req": "UC 9", "segment": "Dashboard Frontend Next.js",
        "id": "TC-UC9-06",
        "nom": "Deconnexion et redirection vers login",
        "criticite": "Moyenne",
        "prerequis": "Medecin connecte au dashboard",
        "etapes": "1. Cliquer sur le bouton de deconnexion\n2. Verifier la suppression des tokens\n3. Verifier la redirection vers la page de connexion",
        "resultat": "Tokens supprimes et redirection vers /login",
        "etat": "Reussi",
        "observations": "Session completement nettoyee",
        "solution": "-",
        "evidences": "sc_08_deconnexion.png",
    },
    # UC 10 : App Mobile Health Connect (10 cas)
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-01",
        "nom": "Ecran de login de l'application mobile",
        "criticite": "Haute",
        "prerequis": "Preview mobile accessible, backend demarre",
        "etapes": "1. Ouvrir l'application mobile\n2. Verifier la presence du formulaire de connexion\n3. Verifier les champs email et mot de passe",
        "resultat": "Formulaire de login affiche avec logo Mood-IoT et champs de saisie",
        "etat": "Reussi",
        "observations": "Interface conforme aux specifications",
        "solution": "-",
        "evidences": "sc_mob_01_login.png",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-02",
        "nom": "Connexion patient avec JWT (spinner)",
        "criticite": "Haute",
        "prerequis": "Identifiants patient valides (sophie.dupont@email.fr)",
        "etapes": "1. Saisir email et mot de passe du patient\n2. Cliquer sur SE CONNECTER\n3. Verifier le spinner de chargement\n4. Verifier la requete POST /auth/login dans la console",
        "resultat": "Spinner affiche, requete HTTP envoyee, token JWT recu",
        "etat": "Reussi",
        "observations": "Token JWT stocke en memoire pour les requetes suivantes",
        "solution": "-",
        "evidences": "sc_mob_02_login_loading.png",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-03",
        "nom": "Accueil patient apres authentification",
        "criticite": "Haute",
        "prerequis": "Patient authentifie",
        "etapes": "1. Verifier l'ecran d'accueil apres login\n2. Verifier le message de bienvenue avec le nom du patient\n3. Verifier les statistiques vides avant synchronisation",
        "resultat": "Ecran principal affiche avec 'Bonjour, Sophie' et stats a '--'",
        "etat": "Reussi",
        "observations": "Nom du patient recupere depuis la reponse du backend",
        "solution": "-",
        "evidences": "sc_mob_03_accueil.png",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-04",
        "nom": "Synchronisation automatique Health Connect",
        "criticite": "Haute",
        "prerequis": "Patient connecte, Health Connect avec donnees",
        "etapes": "1. Observer la synchronisation automatique apres login\n2. Verifier la lecture Health Connect (Steps, HeartRate, SleepSession)\n3. Verifier le spinner de synchronisation",
        "resultat": "Lecture automatique des capteurs et envoi au backend lance",
        "etat": "Reussi",
        "observations": "Sync automatique 1 seconde apres le login",
        "solution": "-",
        "evidences": "sc_mob_04_sync_loading.png",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-05",
        "nom": "Donnees synchronisees avec succes (badge vert)",
        "criticite": "Haute",
        "prerequis": "Synchronisation en cours",
        "etapes": "1. Attendre la fin de la synchronisation\n2. Verifier le badge vert 'Synchronise a HH:MM'\n3. Verifier les valeurs PAS, BPM, SOMMEIL affichees\n4. Verifier la reponse du serveur dans la console",
        "resultat": "Badge vert, donnees affichees, reponse upserted:true dans la console",
        "etat": "Reussi",
        "observations": "Console affiche le payload envoye et la reponse du serveur",
        "solution": "-",
        "evidences": "sc_mob_05_sync_ok.png",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-06",
        "nom": "Re-synchronisation UPSERT (pas de doublon)",
        "criticite": "Haute",
        "prerequis": "Premiere synchronisation effectuee",
        "etapes": "1. Cliquer sur SYNCHRONISER une deuxieme fois\n2. Verifier que les nouvelles donnees sont envoyees\n3. Verifier dans la console que l'UPSERT met a jour sans creer de doublon",
        "resultat": "Donnees mises a jour, une seule ligne par (patient_id, date) en BD",
        "etat": "Reussi",
        "observations": "UPSERT ON CONFLICT (patient_id, date) DO UPDATE",
        "solution": "-",
        "evidences": "sc_mob_06_resync.png",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-07",
        "nom": "Verification donnees dans PostgreSQL",
        "criticite": "Haute",
        "prerequis": "Synchronisation effectuee",
        "etapes": "1. Executer SELECT sur daily_aggregates pour le patient\n2. Verifier heart_rate_avg, step_count, sleep_duration_min\n3. Verifier source_platform = android_health_connect\n4. Verifier synced_at horodate",
        "resultat": "Donnees presentes dans PostgreSQL avec tous les champs remplis",
        "etat": "Reussi",
        "observations": "Champ source_platform permet de distinguer Android vs iOS",
        "solution": "-",
        "evidences": "sc_mob_07_db_check.txt",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-08",
        "nom": "Batch sync de 3 jours offline",
        "criticite": "Haute",
        "prerequis": "Token JWT valide, endpoint batch disponible",
        "etapes": "1. POST /patients/{id}/health-data/batch avec 3 jours de donnees\n2. Verifier la reponse synced_count = 3\n3. Verifier les 3 enregistrements en BD",
        "resultat": "3 jours synchronises en une seule requete, reponse synced_count:3",
        "etat": "Reussi",
        "observations": "Simule un patient qui n'a pas ouvert l'app pendant 3 jours",
        "solution": "-",
        "evidences": "sc_mob_08_batch_api.json",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-09",
        "nom": "Verification batch dans PostgreSQL",
        "criticite": "Haute",
        "prerequis": "Batch sync effectue",
        "etapes": "1. SELECT COUNT(*) FROM daily_aggregates pour le patient\n2. Verifier que tous les jours sont presents\n3. Verifier l'absence de doublons",
        "resultat": "7 enregistrements uniques dans daily_aggregates",
        "etat": "Reussi",
        "observations": "Chaque jour = une seule ligne grace a l'UPSERT",
        "solution": "-",
        "evidences": "sc_mob_09_db_batch.txt",
    },
    {
        "req": "UC 10", "segment": "App Mobile Health Connect",
        "id": "TC-UC10-10",
        "nom": "Deconnexion de l'application mobile",
        "criticite": "Moyenne",
        "prerequis": "Patient connecte",
        "etapes": "1. Cliquer sur Deconnexion\n2. Verifier la suppression du token JWT\n3. Verifier le retour a l'ecran de login",
        "resultat": "Token supprime, retour a l'ecran de connexion",
        "etat": "Reussi",
        "observations": "Aucune donnee sensible accessible apres deconnexion",
        "solution": "-",
        "evidences": "sc_mob_10_logout.png",
    },
]

# ---------------------------------------------------------------------------
# Headers de la feuille Tests
# ---------------------------------------------------------------------------
HEADERS_TESTS = [
    "# Req. Fonctionnel", "Segment de tests", "Numero de Cas",
    "Nom du cas de test", "Criticite", "Pre-requis", "Etapes",
    "Resultat attendu", "Etat du cas", "Observations",
    "Solution Cas Echoues", "Evidences",
]
COL_WIDTHS_TESTS = [16, 30, 14, 40, 12, 34, 46, 40, 14, 32, 22, 44]


# ---------------------------------------------------------------------------
# Fonctions utilitaires
# ---------------------------------------------------------------------------
def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


# ---------------------------------------------------------------------------
# Feuille 1 : Historique Doc.
# ---------------------------------------------------------------------------
def create_historique(wb):
    ws = wb.active
    ws.title = "Historique Doc."
    ws.sheet_properties.tabColor = BLEU_FONCE

    # Titre
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "GUIDE DE TESTS TECHNIQUES / FONCTIONNELS"
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    ws.merge_cells("F1:H1")
    ws["F1"].value = "IDENTIFICATION : MI.QA.2026-04-12"
    ws["F1"].font = Font(name="Calibri", bold=True, size=10, color=BLEU_FONCE)

    info = [
        ("Nom Document", "Guide de Tests Fonctionnels - Mood-IoT"),
        ("Service", "Backend (FastAPI) + Frontend (Next.js) + PostgreSQL"),
        ("Developpeur", "Cinthya Basilio"),
        ("Certificateur", "Equipe QA"),
        ("Date", "12/04/2026"),
        ("Version", "1.0"),
        ("Nombre de cas de test", str(len(TEST_CASES))),
    ]

    for i, (label, val) in enumerate(info, start=3):
        cell_l = ws.cell(row=i, column=1, value=label)
        cell_l.font = Font(name="Calibri", bold=True, size=10)
        cell_l.border = BORDER_THIN
        cell_l.fill = FILL_BLEU_CLAIR
        cell_v = ws.cell(row=i, column=2, value=val)
        cell_v.font = FONT_BODY
        cell_v.border = BORDER_THIN

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    # Historique des versions
    row = len(info) + 5
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Historique des versions").font = Font(
        name="Calibri", bold=True, size=12, color=BLEU_FONCE
    )
    row += 1
    ver_headers = ["Version", "Date", "Auteur", "Description"]
    for j, h in enumerate(ver_headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER_THIN
        c.alignment = ALIGN_CENTER
    row += 1
    ver_data = ["1.0", "12/04/2026", "Cinthya Basilio", "Creation initiale du guide de tests"]
    for j, v in enumerate(ver_data, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = FONT_BODY
        c.border = BORDER_THIN
        c.alignment = ALIGN_CENTER


# ---------------------------------------------------------------------------
# Feuille 2 : Tests
# ---------------------------------------------------------------------------
def create_tests(wb):
    ws = wb.create_sheet("Tests")
    ws.sheet_properties.tabColor = BLEU_MOYEN

    # Titre
    ws["A1"].value = "GUIDE DE TESTS - FONCTIONNELS"
    ws["A1"].font = Font(bold=True, size=13, color=BLEU_FONCE)
    ws["A2"].value = "PROJET :"
    ws["A2"].font = Font(bold=True, size=10)
    ws["B2"].value = "Guide de Tests Fonctionnels - Mood-IoT"
    ws["F2"].value = "Phase du projet :"
    ws["F2"].font = Font(bold=True, size=10)
    ws["G2"].value = "Developpement"

    # Compteurs en haut a droite
    total = len(TEST_CASES)
    reussis = sum(1 for tc in TEST_CASES if tc["etat"] == "Reussi")
    echoues = sum(1 for tc in TEST_CASES if tc["etat"] == "Echoue")

    ws["J2"].value = "REUSSIS"
    ws["J2"].font = Font(bold=True, color="006100")
    ws["K2"].value = "ECHOUES"
    ws["K2"].font = Font(bold=True, color="9C0006")
    ws["L2"].value = "TOTAL"
    ws["L2"].font = Font(bold=True)
    ws["J3"].value = reussis
    ws["J3"].fill = FILL_GREEN
    ws["K3"].value = echoues
    ws["K3"].fill = FILL_GREEN if echoues == 0 else FILL_RED
    ws["L3"].value = total

    # En-tetes
    for j, h in enumerate(HEADERS_TESTS, 1):
        ws.cell(row=4, column=j, value=h)
    style_header_row(ws, 4, 12)

    ws.auto_filter.ref = f"A4:L{4 + len(TEST_CASES)}"
    ws.freeze_panes = "A5"

    for i, tc in enumerate(TEST_CASES, start=5):
        vals = [
            tc["req"], tc["segment"], tc["id"], tc["nom"],
            tc["criticite"], tc["prerequis"], tc["etapes"],
            tc["resultat"], tc["etat"], tc["observations"],
            tc["solution"], tc["evidences"],
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = FONT_BODY
            c.border = BORDER_THIN
            c.alignment = ALIGN_LEFT

        # Couleur Etat du cas
        etat_cell = ws.cell(row=i, column=9)
        etat_cell.alignment = ALIGN_CENTER
        if tc["etat"] == "Reussi":
            etat_cell.fill = FILL_GREEN
            etat_cell.font = Font(name="Calibri", bold=True, size=10, color="006100")
        elif tc["etat"] == "Echoue":
            etat_cell.fill = FILL_RED
            etat_cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
        elif tc["etat"] == "En attente":
            etat_cell.fill = FILL_YELLOW
            etat_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")

        # Couleur criticite
        crit_cell = ws.cell(row=i, column=5)
        crit_cell.alignment = ALIGN_CENTER
        if tc["criticite"] == "Haute":
            crit_cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
        elif tc["criticite"] == "Moyenne":
            crit_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")

        # Alternance lignes
        if i % 2 == 0:
            for j in range(1, 13):
                cell = ws.cell(row=i, column=j)
                if cell.fill == PatternFill():
                    cell.fill = FILL_ALT

    # Largeurs de colonnes
    for j, w in enumerate(COL_WIDTHS_TESTS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ---------------------------------------------------------------------------
# Feuille 3 : Avancement FINAL
# ---------------------------------------------------------------------------
def create_avancement(wb):
    ws = wb.create_sheet("Avancement FINAL")
    ws.sheet_properties.tabColor = "00B050"

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"].value = "AVANCEMENT DES TESTS FONCTIONNELS - MOOD-IOT"
    ws["A1"].font = Font(bold=True, size=13, color=BLEU_FONCE)

    total = len(TEST_CASES)
    reussis = sum(1 for tc in TEST_CASES if tc["etat"] == "Reussi")
    echoues = sum(1 for tc in TEST_CASES if tc["etat"] == "Echoue")
    en_attente = sum(1 for tc in TEST_CASES if tc["etat"] == "En attente")

    resume = [
        ("Total des cas de test", total),
        ("Cas de test reussis", reussis),
        ("Cas de test echoues", echoues),
        ("Cas de test en attente", en_attente),
        ("Pourcentage d'avancement general", f"{round(reussis / total * 100)}%"),
    ]
    for i, (k, v) in enumerate(resume, 3):
        cell_k = ws.cell(row=i, column=1, value=k)
        cell_k.font = Font(name="Calibri", bold=True, size=10)
        cell_k.border = BORDER_THIN
        cell_k.fill = FILL_BLEU_CLAIR
        cell_v = ws.cell(row=i, column=2, value=v)
        cell_v.font = Font(name="Calibri", bold=True, size=10)
        cell_v.border = BORDER_THIN
        cell_v.alignment = ALIGN_CENTER
        if k == "Cas de test reussis":
            cell_v.fill = FILL_GREEN
        elif k == "Cas de test echoues" and echoues > 0:
            cell_v.fill = FILL_RED

    # Tableau par UC
    row = 10
    ws.cell(row=row, column=1, value="Ventilation par cas d'utilisation").font = Font(
        bold=True, size=11, color=BLEU_FONCE
    )
    row += 1
    headers = ["Cas d'utilisation", "Segment", "Total cas", "Reussis", "Echoues", "Taux de reussite"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER_THIN
        c.alignment = ALIGN_CENTER

    uc_data = {}
    for tc in TEST_CASES:
        key = tc["req"]
        if key not in uc_data:
            uc_data[key] = {"segment": tc["segment"], "total": 0, "reussis": 0, "echoues": 0}
        uc_data[key]["total"] += 1
        if tc["etat"] == "Reussi":
            uc_data[key]["reussis"] += 1
        elif tc["etat"] == "Echoue":
            uc_data[key]["echoues"] += 1

    row += 1
    for uc, d in uc_data.items():
        taux = f"{(d['reussis'] / d['total'] * 100):.0f}%"
        vals = [uc, d["segment"], d["total"], d["reussis"], d["echoues"], taux]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.font = FONT_BODY
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER
        ws.cell(row=row, column=6).fill = FILL_GREEN
        ws.cell(row=row, column=6).font = Font(name="Calibri", bold=True, size=10, color="006100")
        row += 1

    # Total general
    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value="TOTAL GENERAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=1).border = BORDER_THIN
    ws.cell(row=row, column=2).border = BORDER_THIN
    for j, v in enumerate([None, None, total, reussis, echoues, f"{round(reussis / total * 100)}%"], 1):
        if v is not None:
            c = ws.cell(row=row, column=j, value=v)
            c.font = Font(name="Calibri", bold=True, size=11)
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER
    ws.cell(row=row, column=6).fill = FILL_GREEN
    ws.cell(row=row, column=6).font = Font(name="Calibri", bold=True, size=11, color="006100")

    for j, w in enumerate([20, 36, 12, 12, 12, 20], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ---------------------------------------------------------------------------
# Feuille 4 : Evidences
# ---------------------------------------------------------------------------
def create_evidences(wb):
    ws = wb.create_sheet("Evidences")
    ws.sheet_properties.tabColor = "FFC000"

    ws.merge_cells("A1:C1")
    ws["A1"].value = "EVIDENCES DES TESTS FONCTIONNELS - MOOD-IOT"
    ws["A1"].font = Font(bold=True, size=13, color=BLEU_FONCE)

    headers = ["Numero de Cas", "Nom du cas de test", "Fichier(s) d'evidence"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    style_header_row(ws, 2, 3)
    ws.freeze_panes = "A3"

    for i, tc in enumerate(TEST_CASES, start=3):
        for j, v in enumerate([tc["id"], tc["nom"], tc["evidences"]], 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = FONT_BODY
            c.border = BORDER_THIN
            c.alignment = ALIGN_LEFT
        if i % 2 == 0:
            for j in range(1, 4):
                ws.cell(row=i, column=j).fill = FILL_ALT

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 58


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    create_historique(wb)
    create_tests(wb)
    create_avancement(wb)
    create_evidences(wb)
    wb.save(OUTPUT)
    print(f"[OK] Fichier Excel genere : {OUTPUT}")
    print(f"     {len(TEST_CASES)} cas de test sur 4 feuilles.")


if __name__ == "__main__":
    main()
